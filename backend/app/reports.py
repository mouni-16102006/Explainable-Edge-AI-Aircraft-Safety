import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_pdf_report(flight_data: dict, prediction_results: dict) -> bytes:
    """
    Generates a beautifully structured PDF safety audit report for aviation operations.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles for Premium Aerospace Look
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#0F172A'), # Slate 900
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#64748B'), # Slate 500
        spaceAfter=20
    )
    
    section_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#1E3A8A'), # Navy 900
        spaceBefore=12,
        spaceAfter=8,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#334155') # Slate 700
    )
    
    rec_style = ParagraphStyle(
        'RecommendationText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=15,
        textColor=colors.HexColor('#991B1B') if prediction_results['status_color'] in ['red', 'orange'] else colors.HexColor('#065F46')
    )

    elements = []
    
    # Header Banner
    elements.append(Paragraph("✈️ AEROSENTINEL SAFETY REPORT", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | System: Edge AI Predictor v1.4", subtitle_style))
    elements.append(Spacer(1, 10))
    
    # Overview Table
    safety_state = prediction_results.get("safety_status", "NORMAL")
    risk_score = prediction_results.get("risk_score", 0.0) * 100
    confidence = prediction_results.get("confidence", 0.0) * 100
    
    summary_data = [
        [Paragraph("<b>Flight Safety Status:</b>", body_style), Paragraph(f"<b>{safety_state}</b>", rec_style)],
        [Paragraph("<b>Predicted Risk Score:</b>", body_style), Paragraph(f"{risk_score:.1f}%", body_style)],
        [Paragraph("<b>Model Confidence:</b>", body_style), Paragraph(f"{confidence:.1f}%", body_style)],
        [Paragraph("<b>Timestamp:</b>", body_style), Paragraph(datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC'), body_style)]
    ]
    
    summary_table = Table(summary_data, colWidths=[180, 320])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 15))
    
    # System Recommendation
    elements.append(Paragraph("AI DECISION SUPPORT RECOMMENDATION", section_style))
    rec_box_data = [[Paragraph(prediction_results.get("recommendation", "N/A"), body_style)]]
    rec_table = Table(rec_box_data, colWidths=[500])
    # Background color based on threat level
    bg_color = '#FEF2F2' if safety_state in ['CRITICAL', 'EMERGENCY'] else ('#FEF3C7' if safety_state == 'WARNING' else '#ECFDF5')
    border_color = '#F87171' if safety_state in ['CRITICAL', 'EMERGENCY'] else ('#FBBF24' if safety_state == 'WARNING' else '#34D399')
    
    rec_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor(bg_color)),
        ('BOX', (0,0), (-1,-1), 1.5, colors.HexColor(border_color)),
        ('PADDING', (0,0), (-1,-1), 12),
    ]))
    elements.append(rec_table)
    elements.append(Spacer(1, 15))
    
    # Failure Probability Table
    elements.append(Paragraph("SUBSYSTEM FAILURE PROBABILITIES", section_style))
    prob_data = [["Subsystem", "Risk Probability", "Status"]]
    for component, prob in prediction_results.get("probabilities", {}).items():
        comp_name = component.replace('_', ' ').replace('fault', '').strip().title()
        status_str = "Nominal"
        if prob > 0.70:
            status_str = "Emergency"
        elif prob > 0.40:
            status_str = "Critical"
        elif prob > 0.15:
            status_str = "Warning"
            
        prob_data.append([comp_name, f"{prob * 100:.1f}%", status_str])
        
    prob_table = Table(prob_data, colWidths=[200, 150, 150])
    prob_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F1F5F9')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'), # Left-align first column
    ]))
    elements.append(prob_table)
    elements.append(Spacer(1, 15))
    
    # Telemetry Audit Table
    elements.append(Paragraph("FLIGHT TELEMETRY AUDIT RECORD", section_style))
    telemetry_data = [["Sensor Parameter", "Measured Value", "Operational Limit"]]
    limits = {
        "engine_temp": "70 - 120 C", "oil_pressure": "40 - 70 PSI", "hydraulic_pressure": "2600 - 3400 PSI",
        "fuel_flow": "1500 - 3500 kg/h", "fuel_pressure": "30 - 50 PSI", "vibration": "1.0 - 6.0 mm/s",
        "rpm": "7000 - 10000 RPM", "voltage": "26 - 30 V", "current": "90 - 150 A",
        "battery_soc": "70 - 100%", "altitude": "0 - 45000 ft", "speed": "150 - 550 kts",
        "cabin_pressure": "10 - 13 PSI", "cabin_temp": "18 - 26 C", "wind_speed": "0 - 50 kts"
    }
    
    for sensor, val in flight_data.items():
        if sensor in limits:
            sensor_name = sensor.replace('_', ' ').title()
            telemetry_data.append([sensor_name, f"{val:.2f}", limits[sensor]])
            
    telemetry_table = Table(telemetry_data, colWidths=[200, 150, 150])
    telemetry_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#334155')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 5),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
    ]))
    elements.append(telemetry_table)
    
    # Build Document
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_excel_report(flight_data: dict, prediction_results: dict) -> bytes:
    """
    Generates a professionally styled Excel workbook containing safety assessments and telemetry logs.
    """
    wb = Workbook()
    
    # Sheet 1: Safety Summary
    ws1 = wb.active
    ws1.title = "Safety Assessment"
    ws1.views.sheetView[0].showGridLines = True
    
    # Styling Helpers
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    light_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    gray_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Header Banner
    ws1.merge_cells("A1:D1")
    ws1["A1"] = "✈️ AEROSENTINEL EDGE AI SAFETY ASSESSMENT"
    ws1["A1"].font = font_title
    ws1["A1"].fill = navy_fill
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40
    
    # Summary Details
    ws1["A3"] = "Safety Status:"
    ws1["A3"].font = font_bold
    ws1["B3"] = prediction_results.get("safety_status", "NORMAL")
    ws1["B3"].font = font_bold
    
    ws1["A4"] = "Risk Index:"
    ws1["A4"].font = font_bold
    ws1["B4"] = f"{prediction_results.get('risk_score', 0.0)*100:.2f}%"
    ws1["B4"].font = font_regular
    
    ws1["A5"] = "AI Confidence:"
    ws1["A5"].font = font_bold
    ws1["B5"] = f"{prediction_results.get('confidence', 0.0)*100:.2f}%"
    ws1["B5"].font = font_regular
    
    ws1["A6"] = "Timestamp:"
    ws1["A6"].font = font_bold
    ws1["B6"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')
    ws1["B6"].font = font_regular
    
    # Subsystem Failure probabilities
    ws1["A8"] = "Subsystem Component"
    ws1["B8"] = "Risk Probability"
    ws1["C8"] = "Risk Level"
    for col in ["A8", "B8", "C8"]:
        ws1[col].font = font_bold
        ws1[col].fill = gray_fill
        ws1[col].border = thin_border
        
    row_idx = 9
    for comp, prob in prediction_results.get("probabilities", {}).items():
        comp_name = comp.replace('_', ' ').replace('fault', '').strip().title()
        status_str = "Nominal"
        if prob > 0.70:
            status_str = "Emergency"
        elif prob > 0.40:
            status_str = "Critical"
        elif prob > 0.15:
            status_str = "Warning"
            
        ws1.cell(row=row_idx, column=1, value=comp_name).font = font_regular
        ws1.cell(row=row_idx, column=2, value=f"{prob*100:.2f}%").font = font_regular
        ws1.cell(row=row_idx, column=3, value=status_str).font = font_regular
        
        for col_idx in range(1, 4):
            ws1.cell(row=row_idx, column=col_idx).border = thin_border
        row_idx += 1
        
    # Recommendation Box
    ws1.cell(row=row_idx+1, column=1, value="AI Maintenance Action Plan:").font = font_section
    ws1.merge_cells(start_row=row_idx+2, start_column=1, end_row=row_idx+4, end_column=4)
    rec_cell = ws1.cell(row=row_idx+2, column=1, value=prediction_results.get("recommendation", "N/A"))
    rec_cell.font = font_bold
    rec_cell.alignment = Alignment(wrap_text=True, vertical="top")
    rec_cell.fill = light_blue_fill
    
    # Sheet 2: Telemetry Data
    ws2 = wb.create_sheet(title="Telemetry Log")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "Sensor Parameter"
    ws2["B1"] = "Measured Value"
    ws2["C1"] = "Engineering Units"
    for col in ["A1", "B1", "C1"]:
        ws2[col].font = font_bold
        ws2[col].fill = gray_fill
        ws2[col].border = thin_border
        
    t_row = 2
    units = {
        "engine_temp": "Celsius", "oil_pressure": "PSI", "hydraulic_pressure": "PSI",
        "fuel_flow": "kg/hour", "fuel_pressure": "PSI", "vibration": "mm/sec",
        "rpm": "RPM", "voltage": "Volts", "current": "Amperes",
        "battery_soc": "Percent (%)", "altitude": "Feet (ft)", "speed": "Knots (kts)",
        "cabin_pressure": "PSI", "cabin_temp": "Celsius", "wind_speed": "Knots (kts)"
    }
    
    for sens, val in flight_data.items():
        if sens in units:
            ws2.cell(row=t_row, column=1, value=sens.replace('_', ' ').title()).font = font_regular
            ws2.cell(row=t_row, column=2, value=float(val)).font = font_regular
            ws2.cell(row=t_row, column=3, value=units[sens]).font = font_regular
            for col_idx in range(1, 4):
                ws2.cell(row=t_row, column=col_idx).border = thin_border
            t_row += 1
            
    # Auto-fit columns
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes

def generate_csv_report(flight_data: dict, prediction_results: dict) -> str:
    """
    Generates a comma-separated CSV summary string containing flight status and sensor values.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write summary metadata
    writer.writerow(["--- AEROSENTINEL FLIGHT SAFETY AUDIT ---"])
    writer.writerow([])
    writer.writerow(["Metric", "Value", "Status"])
    writer.writerow(["Overall Safety Status", prediction_results.get("safety_status", "NORMAL")])
    writer.writerow(["Risk Index Score (%)", f"{prediction_results.get('risk_score', 0.0)*100:.2f}"])
    writer.writerow(["Prediction Confidence (%)", f"{prediction_results.get('confidence', 0.0)*100:.2f}"])
    writer.writerow(["Timestamp", datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')])
    writer.writerow([])
    
    # Write component risks
    writer.writerow(["--- SUBSYSTEM FAILURES ---"])
    writer.writerow(["Subsystem Component", "Failure Probability (%)"])
    for comp, prob in prediction_results.get("probabilities", {}).items():
        comp_name = comp.replace('_', ' ').replace('fault', '').strip().title()
        writer.writerow([comp_name, f"{prob*100:.2f}"])
    writer.writerow([])
    
    # Write sensor values
    writer.writerow(["--- TELEMETRY READINGS ---"])
    writer.writerow(["Sensor Name", "Measured Value"])
    for sens, val in flight_data.items():
        writer.writerow([sens.replace('_', ' ').title(), f"{val:.4f}"])
        
    csv_str = output.getvalue()
    output.close()
    return csv_str
